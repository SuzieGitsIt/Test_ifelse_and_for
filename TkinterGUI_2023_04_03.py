# File:     TkinterGUI_2023-04-03
# Version:  0.0.01
# Author:   Susan Haynes
# Comments/Notes: 
#   (0,0) coordinates are the top left corner of the screen for 1920x1080
#   (0,0) coordinates are the bottom right corner of the screen for 1919x1079
# Online References: 
#   https://pypi.org/project/PyAutoGUI/
#   https://pyautoGUI.readthedocs.io/en/latest/mouse.html
# Revision History: N/A 
# To check tkinter is installed, use this in command promt.
# python -m tkinter 

#################################################      CLASSES OF LIBRARIES TO USE      ################################################  
from functools import partial                           # for allowing 015 040 buttons to equal specific values when clicked.
from openpyxl import *                                  # Write to excel
import xlsxwriter                                       # Excel Writer library 
import tkinter as tk                                    # Tkinter's Tk class
import tkinter.ttk as ttk                               # Tkinter's Tkk class
import datetime as dt                                   # Date library
import pandas as pd
from PIL import ImageTk, Image                          # Displaying LAL background photo
from tkinter import messagebox                          # Exit standard message box

##########################################################################################################################################
#################################################         1st    GUI SCREEN              #################################################
##########################################################################################################################################
#################################################      INITIALIZING STANDARD DISPLAY     ################################################# 
GUI = tk.Tk()
GUI.title("LAL Measurement")
GUI.geometry("1000x700")                                # Set the geometry of Tkinter frame
GUI.configure(bg = 'white')                             # Set background color
GUI.option_add("*Font", "Helvetica 12 bold")            # set the font and size for entire GUI
GUI.option_add("*fg", "black")                          # set the text color, hex works too "#FFFFFF"
GUI.option_add("*bg", "white")                          # set the background to white

#################################################            BUTTON PRESS STYLE           ################################################ 
style = ttk.Style(); 
style.theme_use('default');     # alt, default, clam and classic
#style.configure('T.Button', font=('Helvetica' 12, weight= "bold")
#style.configure('P.Button', font=('Helvetica' 12, weight= "bold")
#style.configure('T.Button',highlightthickness='10')             
#style.configure('W.Button',highlightcolor='pink')
style.map('T.TButton',background=[('active', 'pressed', 'white'),('!active','white'), ('active','!pressed','grey')]) # active, not active, not pressed
style.map('T.TButton',relief    =[('pressed','sunken'),('!pressed','raised')]) # pressed, not pressed
style.configure("T.Button", font= ('Helvetica', '12', 'bold'))

style.map('P.TButton',background=[('active', 'pressed', '#FF69B4'),('!active','white'), ('active','!pressed','grey')]) # Press me Button always hot pink when pressed
style.map('P.TButton',relief    =[('pressed','sunken'),('!pressed','raised')]) # pressed, not pressed
style.configure("P.Button", font= ('Helvetica', '12', 'bold'))

# Python is serial, so each widget will output in the order listed below;
#################################################           LAL BACKGROUND IMAGE          ################################################  
def resize_image(event):
    new_width = event.width
    new_height = event.height
    background_image = copy_of_image.resize((new_width, new_height))
    bkgnd_img = ImageTk.PhotoImage(background_image)
    lbl_photo.config(image = bkgnd_img)
    lbl_photo.background_image = bkgnd_img #avoid garbage collection

background_image = Image.open(r"\\RXS-FS-02\userdocs\shaynes\My Documents\R&D - Software\Python\TkinterGUI_2023-03-24\LAL.png")
copy_of_image = background_image.copy()
bkgnd_img = ImageTk.PhotoImage(background_image)

lbl_photo = ttk.Label(GUI, image = bkgnd_img)
lbl_photo.bind('<Configure>', resize_image)
lbl_photo.pack(fill=tk.BOTH, expand = True)

################################################         DISPLAY TODAY'S DATE            ################################################   
date = dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
lbl_cmd_date = tk.Label(text="Today's Date is:", bg= "white", width= 12).place(x=50,y=50)  
lbl_out_date = tk.Label(GUI, text=f"{dt.datetime.now():%m-%d-%Y}", bg= "white", width= 9).place(x=300, y=50) 

#################################################               EXCEL FILE               ################################################  
#filename = 
try:
    workbook = load_workbook(f"{filename},{date}.xlsx", index=False)
    sheet = workbook.active
except:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Credentials"
    sheet["B1"] = "Work Order"
    sheet["C1"] = "Sample Size"
    sheet["D1"] = "Measurement Size"
    sheet["E1"] = "Date and Time"

new_line = sheet.max_row + 1

################################################         NEW EXCIL FILE NAME              ################################################   

#lbl_out_new_excel = (GUI, label_out_cred)
#lbl_out_new_excel.place(x=300, y=100) 

################################################                 MAIN BODY                ################################################   
# Display the command label before the entry box to indicate what information the Opterator is to type
lbl_cmd_cred = tk.Label(text="Enter Operator Credentials:", bg= "white", width= 20).place(x=50,y=150) 
lbl_cmd_WO   = tk.Label(text="Enter Work Order Number:",    bg= "white", width= 20).place(x=50,y=200)   
lbl_cmd_samp = tk.Label(text="Enter Sample Size:",          bg= "white", width= 14).place(x=50,y=250)  
lbl_cmd_meas = tk.Label(text="Select Measurement Size:",    bg= "white", width= 20).place(x=50,y=300)  

# Entry boxes to take information from operator
entry_cred = tk.Entry(GUI, bg= "white", width= 10) 
entry_cred.focus_set()                              # Places cursor in the first entry box.
entry_cred.place(x=300,y=150) 
entry_WO   = tk.Entry(GUI, bg= "white", width= 10) 
entry_WO.place(x=300,y=200) 
entry_samp = tk.Entry(GUI, bg= "white", width= 10) 
entry_samp.place(x=300,y=250)  

# Display the inputs as outputs
lbl_disp_cred = tk.Label(GUI, text="Credentials:",       bg= "white", width= 9) .place(x=50, y=450) 
lbl_disp_WO   = tk.Label(GUI, text="Work Order Number:", bg= "white", width= 16).place(x=50, y=500) 
lbl_disp_samp = tk.Label(GUI, text="Sample Size:",       bg= "white", width= 10).place(x=50, y=550) 
lbl_disp_meas = tk.Label(GUI, text="Measurement Size:",  bg= "white", width= 15).place(x=50, y=600) 

# Display the user inputs as outputs 
lbl_out_cred = tk.Label(GUI, text= "", bg= "white", width= 3) 
lbl_out_cred.place(x=300, y=450) 
lbl_out_WO   = tk.Label(GUI, text= "", bg= "white", width= 10)
lbl_out_WO.place(x=300, y=500) 
lbl_out_samp = tk.Label(GUI, text= "", bg= "white", width= 3) 
lbl_out_samp.place(x=300, y=550) 
 
# Display user inputs
def display_cred(): 
   global entry 
   cred = entry_cred.get()[:3]                          # entry_cred is the variable we are passing. Limit 3 characters
   lbl_out_cred.configure(text = cred)
   print(entry_cred.get()[:3])                          # Print can be removed after developed.

def display_WO():                        
   global entry 
   WO = entry_WO.get()[:10]                             # entry_WO is the variable we are passing. Limit 10 characters
   lbl_out_WO.configure(text = WO) 
   print(entry_WO.get()[:10]) 

def display_samp():                        
   global entry 
   samp = entry_samp.get()[:2]                          # entry_samp is the variable we are passing. Limit 2 characters
   lbl_out_samp.configure(text = samp)
   print(entry_samp.get()[:2]) 

def excel_entry():
    sheet.cell(column=1, row=new_line, value=entry_cred.get()[:3])
    sheet.cell(column=2, row=new_line, value=entry_WO.get()[:10])
    sheet.cell(column=3, row=new_line, value=entry_samp.get()[:2])
    sheet.cell(column=5, row=new_line).value = date
    updater(filename)
    workbook.save(filename="WO+Date+Meas.xlsx") 

def entry_015(text):
    sheet.cell(column=4, row=new_line).value = text
    btn_015 = tk.Entry(GUI, width= 10)
    btn_015.insert(0,text) 
    btn_015.place(x=300, y=600)
    print(text)
    excel_entry()

def entry_040(text):
    sheet.cell(column=4, row=new_line).value = text
    btn_040 = tk.Entry(GUI, width= 10)
    btn_040.insert(0,text) 
    btn_040.place(x=300, y=600)
    print(text)
    excel_entry()

def entry_100(text):
    sheet.cell(column=4, row=new_line).value = text
    btn_100 = tk.Entry(GUI, width= 10)
    btn_100.insert(0,text) 
    btn_100.place(x=300, y=600)
    print(text)
    excel_entry()

#filename = dt.datetime.now().strftime('%Y-%m-%d'), text= f"copy {entry_WO} {entry_samp}"
#print(filename)
 
################################################         BUTTON COUNTER (TOTAL)          ################################################   
btn_pres_cnt = 0                                                         # setting count to 0 to be able to call it a global variable within the function
def pink(event):                     
    global btn_pres_cnt                                                  # initializing btn_pres_cnt as a global varaible so that it adds through every iteration
    if(btn_pres_cnt==5 or btn_pres_cnt==10 or btn_pres_cnt==15 or btn_pres_cnt==20 or btn_pres_cnt==25): # button turns pink when btn_pres_cnt=100, and =200 and = 300.
        style.map('T.TButton',background=[('active', 'pressed', '#FF69B4'),('!active','white'), ('active','!pressed','grey')])    # only the button being pressed turns hot pink
        style.configure("T.Button", font= ('Helvetica', '12', 'bold'))
    else:   # else is the normal style
        style.map('T.TButton',background=[('active', 'pressed', 'white'),('!active','white'), ('active','!pressed','grey')])
        style.configure("T.Button", font= ('Helvetica', '12', 'bold'))
    print("btn_pres_cnt = ", btn_pres_cnt)                               # This is always executed at the end of the if else
    btn_pres_cnt +=1                                                     # This is always executed at the end of the if else

#################################################        BUTTONS TO BE CLICKED         ################################################   
btn_cred = ttk.Button(GUI, text= "Confirm", style= 'T.TButton', command= display_cred)
btn_cred.bind("<Button-1>", pink)
btn_cred.place(x=450,y=140)  
btn_WO   = ttk.Button(GUI, text= "Confirm", style= 'T.TButton', command=display_WO)  
btn_WO.bind("<Button-1>", pink)
btn_WO.place(x=450,y=190)  
btn_samp = ttk.Button(GUI, text= "Confirm", style= 'T.TButton', command=display_samp)
btn_samp.bind("<Button-1>", pink)
btn_samp.place(x=450,y=240)  
btn_015  = ttk.Button(GUI, text="Posterior", style= 'T.TButton', command=partial(entry_015, "Post - 015"))      # Post - 015 is the variable we are passing to excel
btn_015.bind("<Button-1>", pink)
btn_015.place(x=300,y=300)  
btn_040  = ttk.Button(GUI, text="Anterior",  style= 'T.TButton', command=partial(entry_040, "Ant - 040"))       # Ant - 040 is the variable we are passing to excel
btn_040.bind("<Button-1>", pink)
btn_040.place(x=400,y=300) 
btn_100  = ttk.Button(GUI, text="Full Lens", style= 'T.TButton', command=partial(entry_100, "Full - 100"))      # Full - 100 is the variable we are passing to excel
btn_100.bind("<Button-1>", pink)
btn_100.place(x=500,y=300) 
btn_sav = ttk.Button(GUI, text="Save Entries",style= 'T.TButton',command= excel_entry)
btn_sav.bind("<Button-1>", pink)
btn_sav.place(x=650, y=630)

################################################          RANDOM PINK BUTTON           ################################################   
but_pink = ttk.Button(GUI, text= "Next", style= 'P.TButton', width= 5)
but_pink.place(x=800,y=630)

################################################               EXIT BUTTON              ################################################   
def exit_application(): 
    msg_box = tk.messagebox.askquestion('Exit', 'Are you sure you want to exit the application?', icon='warning') 
    if msg_box == 'yes': 
        GUI.destroy() 
    else: 
        tk.messagebox.showinfo('Exit', 'Thanks for staying, please continue.') 

but_exit = ttk.Button(GUI, text="Exit", style= 'T.TButton', width= 5, command=exit_application).place(x=900,y=630) 

# Must be at the end of the program in order for the application to run b/c windows is constantly updating
GUI.mainloop()
